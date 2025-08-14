import itertools
import re
import xlsxwriter
from openpyxl.reader.excel import load_workbook


def add_to_excel (sheetname,headers,row_data,filename='Stock_Management.xlsx'):
    old_data = {}
    try:
        wb_old = load_workbook(filename)
        for sheet in wb_old.sheetnames:
            ws = wb_old[sheet]
            rows = list(ws.iter_rows(values_only=True))
            old_data[sheet] = [list(row) for row in rows]
    except FileNotFoundError:
        old_data = {}

    if sheetname not in old_data:
        old_data[sheetname] = [headers]
    old_data[sheetname].append(row_data)

    workbook = xlsxwriter.Workbook(filename)
    for sheet, data in old_data.items():
        worksheet = workbook.add_worksheet(sheet)
        for r_idx,row in enumerate(data):
            for c_idx,val in enumerate(row):
                worksheet.write(r_idx,c_idx,val)
    workbook.close()

def update_product_sheet(product):
    filename = 'Stock_Management.xlsx'
    try:
        wb = load_workbook(filename)
        ws = wb['Product_master']
        data = [list(row) for row in ws.iter_rows(values_only=True)]
        wb.close()

        if not data:
            print("No Data Found In Product_master Sheet.")
            return

        headers = [str(h).strip().lower() for h in data[0]]
        name_idx = headers.index("name")
        incoming_idx = headers.index("incoming")
        outgoing_idx = headers.index("outgoing")
        on_hand_idx = headers.index("on_hand_stock")

        for row in data[1:]:
            if str(row[name_idx]).strip().lower() == product.name:
                row[incoming_idx] = product.incoming
                row[outgoing_idx] = product.outgoing
                row[on_hand_idx] = product.on_hand_stock
                break

        wb_all = load_workbook(filename)
        all_data = {}
        for sheet in wb_all.sheetnames:
            ws =wb_all[sheet]
            all_data[sheet] = [list(r) for r in ws.iter_rows(values_only=True)]
            all_data["Product_master"] = data
            wb_all.close()

        workbook = xlsxwriter.Workbook(filename)
        for sheet,rows in all_data.items():
            ws = workbook.add_worksheet(sheet)
            for r_idx,row in enumerate(rows):
                for c_idx,val in enumerate(row):
                    ws.write(r_idx, c_idx, val)
        workbook.close()

        print(f"Product '{product.name}' updated in Excel.")

    except Exception as e:
        print("Error Updating product sheet:",e)

def save_data_in_excel(filename = "Stock_Management.xlsx"):
    customer_data = [["Customer_ID","Customer_Name","Pan_No.","Phone_No","Customer_Type"]]
    for cust in Customer_master.customers.values():
        customer_data.append([cust.customer_id, cust.name, cust.pan_no, cust.phone_no, cust.customer_type])

    product_data = [["Product_ID","Product_Name","Incoming","Outgoing","On_Hand_Stock"]]
    for prod in Product_master.products.values():
        product_data.append([prod.product_id, prod.name, prod.incoming, prod.outgoing, prod.on_hand_stock])

    stock_data = [["Transaction_ID","Customer_Name","Product_Name","Quantity","Transaction_Type"]]
    for txn in Stock_master.transactions:
        stock_data.append([txn.transaction_id, txn.customer_name, txn.product_name, txn.quantity, txn.txn_type])

    workbook = xlsxwriter.Workbook(filename)
    for sheetname, data in {
        "Customer_master": customer_data,
        "Product_master": product_data,
        "Stock_master": stock_data
    }.items():
        ws = workbook.add_worksheet(sheetname)
        for r_idx,row in enumerate(data):
            for c_idx,val in enumerate(row):
                ws.write(r_idx,c_idx,val)
    workbook.close()
    print("All Data Saved to Excel on Exit.")

def initialize_from_excel(filename = 'Stock_Management.xlsx'):
    try:
        wb = load_workbook(filename)
        if "Customer_master" in wb.sheetnames:
            rows = list(wb['Customer_master'].iter_rows(values_only=True))[1:]
            max_id = 0
            for row in rows:
                cust_id,name,pan_no,phone_no,cust_type = row
                cust_obj = Customer_master.__new__(Customer_master)
                cust_obj.customer_id = cust_id
                cust_obj.name = name
                cust_obj.pan_no = pan_no
                cust_obj.phone_no = phone_no
                cust_obj.customer_type = cust_type
                Customer_master.customers[pan_no] = cust_obj
                Customer_master.cust_name.add(name)
                Customer_master.phone_numbers.add(phone_no)
                max_id = max(max_id,cust_id)
            Customer_master.customer_id_counter = itertools.count(start=max_id+1)

        if "Product_master" in  wb.sheetnames:
            rows = list(wb['Product_master'].iter_rows(values_only=True))[1:]
            max_id = 0
            for row in rows:
                prod_id,name,incoming,outgoing,on_hand_stock = row
                prod_obj = Product_master.__new__(Product_master)
                prod_obj.product_id = prod_id
                prod_obj.name = name
                prod_obj.incoming = incoming
                prod_obj.outgoing = outgoing
                prod_obj.on_hand_stock = on_hand_stock
                Product_master.products[name] = prod_obj
                max_id = max(max_id,prod_id)
            Product_master.product_id_counter = itertools.count(start=max_id + 1)

        if "Stock_master" in wb.sheetnames:
            rows = list(wb['Stock_master'].iter_rows(values_only=True))[1:]
            max_id = 0
            for row in rows:
                txn_id,cust_name,prod_name,qty,txn_type = row
                txn_obj = Stock_master.__new__(Stock_master)
                txn_obj.transaction_id = txn_id
                txn_obj.customer_name = cust_name
                txn_obj.product_name = prod_name
                txn_obj.quantity = qty
                txn_obj.txn_type = txn_type
                Stock_master.transactions.append(txn_obj)
                max_id = max(max_id, txn_id)
            Stock_master.transaction_id_counter = itertools.count(start = max_id + 1)

    except FileNotFoundError:
        pass


class Customer_master:
    customers = {}
    phone_numbers =set()
    cust_name = set()

    customer_id_counter = itertools.count(start=1)

    def __init__(self,name,pan_no,customer_type,phone_no):
        if not all([name.strip(),pan_no.strip(),customer_type.strip(),phone_no.strip()]):
            raise  ValueError("Required all Customer fields filled!")

        if not re.fullmatch(r"[A-Z]{5}[0-9]{4}[A-Z]",pan_no):
            raise ValueError("Pan No. must be 10 characters: 5 alphabet 4 numeric 1 alphabet!")

        if pan_no in Customer_master.customers:
            raise ValueError("Enter the Unique PAN Number!")

        if customer_type not in ['vendor','customer','both']:
            raise ValueError("Type must be 'Vendor','Customer' or 'Both'!")

        if phone_no in Customer_master.phone_numbers:
            raise ValueError("Phone no must be unique!")

        if not (phone_no.isdigit() and len(phone_no)==10):
            raise ValueError("Enter the 10 Digits Phone No.!")

        if name in Customer_master.cust_name:
            raise ValueError("Customer name must be unique!")

        self.customer_id = next(Customer_master.customer_id_counter)
        self.name = name
        self.pan_no = pan_no
        self.phone_no = phone_no
        self.customer_type = customer_type
        Customer_master.customers[pan_no]=self
        Customer_master.cust_name.add(name)
        Customer_master.phone_numbers.add(phone_no)

        add_to_excel('Customer_master',
                     ['Customer_ID','Customer_Name','Customer_Pan_No.','Customer_phone_No.','Customer_Type'],
                     [self.customer_id,self.name,self.pan_no,self.phone_no,self.customer_type])

    def get_by_name(name):
        for customer in Customer_master.customers.values():
            if customer.name == name:
                return customer
        return None

    def __str__(self):
        return f"{self.customer_id} | Customer_Name:{self.name} | Customer_PAN_no.:{self.pan_no} | Customer_phone_no.:{self.phone_no} | Customer_Type:{self.customer_type}"

class Product_master:
    products = {}
    product_id_counter = itertools.count(start=1)

    def __init__(self,name):

        if not name.strip():
            raise  ValueError("Product name is required!")

        if name in Product_master.products:
            raise ValueError("Enter the Unique Product Name!")

        self.product_id = next(Product_master.product_id_counter)
        self.name = name
        self.incoming = 0
        self.outgoing = 0
        self.on_hand_stock = 0

        Product_master.products[name] = self

        add_to_excel('Product_master',
                     ['Product_ID', 'Product_Name', 'Product_Incoming', 'Product_Outgoing', 'Product_On_Hand_Stock'],
                     [self.product_id, self.name, self.incoming, self.outgoing, self.on_hand_stock])

    def update_stock(self,qty,txn_type):
        if txn_type == "incoming":
            self.incoming += qty
            self.on_hand_stock += qty

        elif txn_type == "outgoing":
            if qty > self.on_hand_stock:
                raise ValueError("Insufficient Stock.")
            self.outgoing += qty
            self.on_hand_stock -= qty

        else:
            raise ValueError("Invalid Transaction Type.")

    def __str__(self):
        return f"{self.product_id} | Product_Name:{self.name} |"f" Incoming:{self.incoming} | Outgoing:{self.outgoing} |" f" On_Hand Stock:{self.on_hand_stock}"

class Stock_master:
    transactions=[]
    transaction_id_counter = itertools.count(start=1)

    def __init__(self,customer_name,product_name,quantity,txn_type):
        customer = Customer_master.get_by_name(customer_name)
        product = Product_master.products.get(product_name)

        if not all([customer_name.strip(),product_name.strip(),str(quantity).strip(),txn_type.strip()]):
            raise  ValueError("Required all transactions fields filled!")

        if not customer:
            raise ValueError("Customer Not Found!")

        if not product:
            raise ValueError("Product Not Found!")
            #product = Product_master(product_name)

        if txn_type not in ['incoming','outgoing']:
            raise ValueError("Invalid Transaction Type!")

        if customer.customer_type == "vendor" and txn_type != "incoming":
            raise ValueError("Vendor can only perform incoming transactions.")

        if customer.customer_type == "customer" and txn_type != "outgoing":
            raise ValueError("Customer can only perform outgoing transactions.")

        self.transaction_id = next(Stock_master.transaction_id_counter)
        self.customer_name = customer_name
        self.product_name = product_name
        self.quantity = quantity
        self.txn_type = txn_type

        product.update_stock(quantity,txn_type)
        Stock_master.transactions.append(self)

        add_to_excel('Stock_master',
                     ['Transaction_ID', 'Customer_Name', 'Product_Name', 'Product_Quantity', 'Transaction_Type'],
                     [self.transaction_id, self.customer_name, self.product_name, self.quantity, self.txn_type])
        update_product_sheet(product)

    @classmethod
    def print_sorted_transaction(cls):
        for t in cls.get_sorted_transaction():
            print(t)

    @classmethod
    def get_sorted_transaction(cls):
        return sorted(cls.transactions, key=lambda t: t.product_name.lower())

    def __str__(self):
        return f"{self.transaction_id} | Customer_Name:{self.customer_name} |Product_Name:{self.product_name} | Quantity:{self.quantity} | Transaction_Type:{self.txn_type}"


def main():
    initialize_from_excel()
    while True:
        print("\nChoose Option:")
        print("1. Add Customer")
        print("2. Add Product")
        print("3. Stock Transaction")
        print("4. Show Customers")
        print("5. Show Product")
        print("6. Show Transaction")
        print("7. Show All Data")
        print("8. Exit")

        choice = input("\nEnter The Choice:")

        try:
            if choice == '1':
                name = input("Enter Customer Name:")
                pan = input("Enter PAN No.:").upper()
                phone_no = input("Enter Phone No.:")
                cust_type = input("Enter Type (Vendor/Customer/Both):").lower()
                cust = Customer_master(name,pan,cust_type,phone_no)
                print("Customer Added:",cust)

            elif choice == '2':
                name = input("Enter Product Name:").lower()
                prod = Product_master(name)
                print("Product Added:",prod)

            elif choice == '3':
                customer_name = input("Enter Customer Name:")
                product_name = input("Enter Product Name:").lower()
                quantity = int(input("Enter Quantity:"))
                txn_type = input("Enter Transaction Type(Incoming/Outgoing):").lower()
                txn = Stock_master(customer_name,product_name,quantity,txn_type)
                print("Transaction Added:",txn)
                print("Updated Product:",Product_master.products[product_name])

            elif choice == '4':
                print("\n----- Customer -----")
                for cust in Customer_master.customers.values():
                    print(cust)

            elif choice == '5':
                print("\n----- Product -----")
                for prod in Product_master.products.values():
                    print(prod)

            elif choice == '6':
                print("\n----- Stock Transactions -----")
                Stock_master.print_sorted_transaction()

            elif choice == '7':
                print("\n----- Customer -----")
                for cust in Customer_master.customers.values():
                    print(cust)

                print("\n----- Product -----")
                for prod in Product_master.products.values():
                    print(prod)

                print("\n----- Stock Transactions -----")
                Stock_master.print_sorted_transaction()

            elif choice == '8':
                save_data_in_excel()
                print("Exiting in Program")
                break

            else:
                print("Invalid Option. Try again!")

        except Exception as e:
            print("Error:",e)


if __name__ == "__main__":
    main()